"""
Exportación a Excel para revisión offline o envío por correo (sección 6,
punto 6 del brief) — equivalente a las hojas que hoy produce Codigo.gs en
Google Sheets, en un único archivo .xlsx multi-hoja.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Caso, Ejecucion, Linea, ProductividadDiaria
from app.services.cruce import calcular_cruce_lote_di
from app.services.reportes import resumen_ejecutivo

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(ws: Worksheet, headers: list[str], rows: list[list], col_widths: list[int] | None = None):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(row)
    widths = col_widths or [18] * len(headers)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fmt(v) -> str:
    return v.strftime("%Y-%m-%d") if v else ""


def generar_excel_auditoria(db: Session) -> bytes:
    wb = Workbook()

    # ---- Resumen ----
    ws = wb.active
    ws.title = "Resumen"
    resumen = resumen_ejecutivo(db)
    _write_sheet(
        ws,
        ["Indicador", "Cantidad"],
        [
            ["Lotes OK (un solo DI)", resumen["lotes_ok"]],
            ["Lotes OK (mismo barco, DI parcial+definitivo)", resumen["lotes_ok_parcial"]],
            ["Lotes a REVISAR (DI de barcos distintos)", resumen["lotes_revisar"]],
            ["Total de lotes en la base de datos", resumen["total_lotes"]],
            ["", ""],
            ["Especie", "Toneladas trazadas"],
            *[[e["especie"], e["toneladas"]] for e in resumen["especies"]],
        ],
        col_widths=[45, 20],
    )

    # ---- Detalle completo ----
    ws = wb.create_sheet("Detalle completo")
    lineas = db.execute(select(Linea).order_by(Linea.fecha_declaracion, Linea.n_declaracion)).scalars().all()
    _write_sheet(
        ws,
        [
            "Lote", "Tipo de Línea", "Estado", "N° Declaración", "DI", "Declaración Origen", "Agente/Barco",
            "Fecha Desembarque DI", "Fecha Elaboración Línea", "Fecha Declaración", "Producto/Recurso", "Especie",
            "Proceso Producto", "Tipo de Frío", "Producto Global", "Preparación Específica", "Código Recurso/Producto",
            "Nombre Bodega", "Toneladas", "Fila original", "Archivo origen",
        ],
        [
            [
                r.lote, r.tipo_linea.value, r.estado.value, r.n_declaracion, r.di, r.declaracion_origen, r.agente,
                _fmt(r.fecha_desembarque_di), _fmt(r.fecha_elaboracion_linea), _fmt(r.fecha_declaracion), r.producto,
                r.especie, r.proceso_producto, r.tipo_frio, r.producto_global, r.preparacion, r.codigo_producto,
                r.nombre_bodega, r.toneladas, r.fila_original, r.archivo_origen,
            ]
            for r in lineas
        ],
        col_widths=[10, 26, 32, 14, 10, 14, 12, 16, 16, 14, 40, 22, 16, 16, 16, 24, 12, 20, 10, 10, 30],
    )

    # ---- Casos a revisar (Historial de Casos) ----
    ws = wb.create_sheet("Casos a revisar")
    casos = db.execute(select(Caso).order_by(Caso.fecha_detectado.desc())).scalars().all()
    _write_sheet(
        ws,
        ["Fecha detectado", "Lote", "DI asociados", "Agentes/Barcos", "N° Declaraciones", "Estado de revisión", "Notas"],
        [
            [c.fecha_detectado.strftime("%Y-%m-%d %H:%M"), c.lote, c.di_asociados, c.agentes, c.n_declaraciones,
             c.estado_revision.value, c.notas]
            for c in casos
        ],
        col_widths=[18, 10, 24, 20, 16, 20, 40],
    )

    # ---- Cruce Lote-DI (3 vías) ----
    ws = wb.create_sheet("Cruce Lote-DI (3 vías)")
    filas_cruce = calcular_cruce_lote_di(db)
    _write_sheet(
        ws,
        ["Lote", "DI externo (registro interno)", "DI resuelto (Sernapesca)", "Estado del cruce", "Estado interno",
         "Especies", "Barco(s)"],
        [
            [f.lote, ", ".join(f.di_externo), ", ".join(f.di_resuelto), f.estado_cruce, f.estado_interno,
             ", ".join(f.especies), ", ".join(f.barcos)]
            for f in filas_cruce
        ],
        col_widths=[10, 24, 24, 22, 20, 22, 18],
    )

    # ---- Productividad Diaria ----
    ws = wb.create_sheet("Productividad Diaria")
    prod = db.execute(select(ProductividadDiaria).order_by(ProductividadDiaria.fecha)).scalars().all()
    _write_sheet(
        ws,
        ["Fecha", "N° Declaraciones", "N° Líneas totales", "N° Líneas Mat.Prima", "N° Líneas Producción",
         "N° DI distintos", "N° Especies distintas", "Toneladas Producto", "Toneladas Materia Prima"],
        [
            [_fmt(p.fecha), p.n_declaraciones, p.n_lineas_totales, p.n_lineas_materia_prima, p.n_lineas_produccion,
             p.n_di_distintos, p.n_especies_distintas, p.toneladas_producto, p.toneladas_materia_prima]
            for p in prod
        ],
        col_widths=[14, 16, 16, 18, 18, 14, 18, 16, 18],
    )

    # ---- Historial de Ejecuciones ----
    ws = wb.create_sheet("Historial de Ejecuciones")
    ejecs = db.execute(select(Ejecucion).order_by(Ejecucion.fecha_ejecucion.desc())).scalars().all()
    _write_sheet(
        ws,
        ["Fecha ejecución", "Archivo procesado", "Lotes OK", "Lotes OK (parcial+definitivo)", "Lotes a revisar",
         "Sin determinar"],
        [
            [e.fecha_ejecucion.strftime("%Y-%m-%d %H:%M"), e.archivo_nombre, e.lotes_ok, e.lotes_ok_parcial,
             e.lotes_revisar, e.lotes_sin_determinar]
            for e in ejecs
        ],
        col_widths=[18, 40, 12, 22, 16, 16],
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
